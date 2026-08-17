"""Fino 机构观点数据层 (移植自 c:/Users/admin/Desktop/fino/run.py).

数据源: https://www.finoview.com.cn/autoApi/foreign/rating_prediction/*
- /overview  机构观点汇总: 各品种看涨/震荡/看跌家数与占比 + 归因文本
- /detail    单品种明细: 各机构逐条观点 (含 uni_id; source 由本地映射表补全)

鉴权: appkey / appsecret, 优先读环境变量 FINO_APPKEY / FINO_APPSECRET,
      未设时回退到模块内置默认 (run.py 原值), 方便本地自托管即开即用.

机构名映射:
  detail API 本身不返回机构名, 只有 uni_id. 与 run.py 一样需要本地映射表:
  uni_id -> 公司名称. 查找顺序:
    1) 环境变量 FINO_UNI_ID_MAP (json / csv / xlsx 路径)
    2) backend/data/fino_uni_id_map.json
    3) backend/data/uni_id_ v1.1_20260119.xlsx (run.py 原文件名)
  json 格式: {"6354": "某某期货", ...} 或 [{"uni_id":"6354","公司名称":"..."}]
  csv/xlsx 需含 uni_id 与 公司名称 两列. 映射缺失时 source 回退为 "机构#<uni_id>".

设计:
- 只读, 无状态, 客观呈现机构观点统计, 不推荐 / 不预测 / 不评分.
- 全站共享一份缓存 (TTL 10 分钟). 过期读上一笔, 不再出网.
- requests 惰性导入: 缺失时抛 DependencyMissing, app 层转 501 + 安装提示.
- 数据源故障的空结果不缓存, 下次请求直接重试.
"""

from __future__ import annotations

import csv
import json
import logging
import os
import time
from pathlib import Path
from typing import Any, Callable

from cache import TTLCache, is_nonempty

logger = logging.getLogger(__name__)

BASE_URL = "https://www.finoview.com.cn"
OVERVIEW_PATH = "/autoApi/foreign/rating_prediction/overview"
DETAIL_PATH = "/autoApi/foreign/rating_prediction/detail"

DEFAULT_HEADERS = {
    "Content-Type": "application/json;charset=UTF-8",
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/139.0.0.0 Safari/537.36"
    ),
}

# 内置默认凭证 (run.py 原值), 公网部署请用环境变量覆盖.
_DEFAULT_APPKEY = "BCEB015C0C0A8187E79B8CE91F7578E6"
_DEFAULT_APPSECRET = "b198f59a281ad4183db165b15a47fb30"

_DATA_DIR = Path(__file__).resolve().parent / "data"
_DEFAULT_MAP_CANDIDATES = (
    _DATA_DIR / "fino_uni_id_map.json",
    _DATA_DIR / "uni_id_ v1.1_20260119.xlsx",
    _DATA_DIR / "uni_id_map.xlsx",
    _DATA_DIR / "uni_id_map.csv",
)


class DependencyMissing(RuntimeError):
    """缺少 requests 依赖时抛出, app 层转 501 + 安装提示."""


def _requests():
    try:
        import requests  # noqa: PLC0415
    except ImportError as e:
        raise DependencyMissing("fino 数据需要 requests: pip install requests") from e
    return requests


_TTL = 600  # 10 分钟, 日报类数据
_CACHE = TTLCache(maxsize=128, default_ttl=_TTL, negative_ttl=0, name="fino")

# uni_id -> company name; None = not loaded yet, {} = loaded but empty/missing file
_UNI_ID_MAP: dict[str, str] | None = None
_UNI_ID_MAP_WARNED = False


def _cached(
    key: str,
    fn,
    valid: Callable[[Any], bool] = is_nonempty,
    ttl: float | None = None,
):
    """First fill then last-good. Empty upstream is not stored, next call retries."""
    return _CACHE.get_or_set(key, fn, ttl=ttl, valid=valid, negative_ttl=0, serve_last=True)


def _creds() -> tuple[str, str]:
    """读取 appkey/appsecret, 优先环境变量, 回退内置默认."""
    appkey = os.environ.get("FINO_APPKEY", "").strip() or _DEFAULT_APPKEY
    appsecret = os.environ.get("FINO_APPSECRET", "").strip() or _DEFAULT_APPSECRET
    return appkey, appsecret


def _post(path: str, payload: dict, timeout: float = 15.0) -> Any:
    """POST JSON, 校验响应壳 code, 返回 data."""
    req = _requests()
    resp = req.post(
        f"{BASE_URL}{path}",
        data=json.dumps(payload),
        headers=DEFAULT_HEADERS,
        verify=False,
        timeout=timeout,
    )
    data = resp.json()
    code = data.get("code")
    if code == 500:
        raise RuntimeError(f"fino 上游错误: {data.get('message') or data}")
    if code != 1:
        raise RuntimeError(f"fino 未知回报码 {code}: {data}")
    return data.get("data") or []


def _norm_date(d: str) -> str:
    """日期归一为 YYYYMMDD; 已是 8 位直接返回, 否则原样返回."""
    s = (d or "").strip()
    if len(s) == 10 and s.count("-") == 2:
        return s.replace("-", "")
    return s


def _map_paths() -> list[Path]:
    """候选映射表路径: 环境变量优先, 其次 backend/data/ 默认名."""
    paths: list[Path] = []
    env = os.environ.get("FINO_UNI_ID_MAP", "").strip()
    if env:
        paths.append(Path(env).expanduser())
    paths.extend(_DEFAULT_MAP_CANDIDATES)
    return paths


def _load_map_from_json(path: Path) -> dict[str, str]:
    raw = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, str] = {}
    if isinstance(raw, dict):
        for k, v in raw.items():
            if k in ("uni_id", "公司名称"):
                continue
            ks, vs = str(k).strip(), str(v).strip() if v is not None else ""
            if ks and vs:
                out[ks] = vs
        # also accept {"mappings": [{"uni_id":..,"公司名称":..}]}
        nested = raw.get("mappings") if isinstance(raw.get("mappings"), list) else None
        if nested:
            for row in nested:
                if not isinstance(row, dict):
                    continue
                uid = str(row.get("uni_id") or "").strip()
                name = str(row.get("公司名称") or row.get("source") or row.get("name") or "").strip()
                if uid and name:
                    out[uid] = name
    elif isinstance(raw, list):
        for row in raw:
            if not isinstance(row, dict):
                continue
            uid = str(row.get("uni_id") or "").strip()
            name = str(row.get("公司名称") or row.get("source") or row.get("name") or "").strip()
            if uid and name:
                out[uid] = name
    return out


def _load_map_from_csv(path: Path) -> dict[str, str]:
    out: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        for row in reader:
            uid = str(row.get("uni_id") or row.get("UNI_ID") or "").strip()
            name = str(
                row.get("公司名称") or row.get("source") or row.get("name") or row.get("company") or ""
            ).strip()
            if uid and name:
                out[uid] = name
    return out


def _load_map_from_xlsx(path: Path) -> dict[str, str]:
    """Read xlsx via pandas (optional) or openpyxl (optional)."""
    try:
        import pandas as pd  # noqa: PLC0415

        df = pd.read_excel(path, dtype={"uni_id": str})
        out: dict[str, str] = {}
        name_col = "公司名称" if "公司名称" in df.columns else None
        if name_col is None:
            for c in ("source", "name", "company"):
                if c in df.columns:
                    name_col = c
                    break
        if "uni_id" not in df.columns or not name_col:
            raise ValueError(f"xlsx missing uni_id/公司名称 columns: {list(df.columns)}")
        for _, row in df.iterrows():
            uid = str(row.get("uni_id") or "").strip()
            name = str(row.get(name_col) or "").strip()
            if uid and name and uid.lower() != "nan" and name.lower() != "nan":
                out[uid] = name
        return out
    except ImportError:
        pass

    try:
        from openpyxl import load_workbook  # noqa: PLC0415
    except ImportError as e:
        raise DependencyMissing(
            "读取 xlsx 映射表需要 pandas 或 openpyxl: pip install pandas openpyxl"
        ) from e

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = [str(c).strip() if c is not None else "" for c in next(rows)]
    try:
        i_uid = header.index("uni_id")
    except ValueError as e:
        raise ValueError(f"xlsx missing uni_id column: {header}") from e
    name_idx = None
    for cand in ("公司名称", "source", "name", "company"):
        if cand in header:
            name_idx = header.index(cand)
            break
    if name_idx is None:
        raise ValueError(f"xlsx missing 公司名称 column: {header}")
    out = {}
    for row in rows:
        if not row or i_uid >= len(row):
            continue
        uid = str(row[i_uid] or "").strip()
        name = str(row[name_idx] or "").strip() if name_idx < len(row) else ""
        if uid and name:
            out[uid] = name
    return out


def _load_uni_id_map() -> dict[str, str]:
    """Load uni_id -> company map.

    Cached after first successful non-empty load. If previously empty (file missing),
    retry on later calls so dropping a map file does not require a full process restart.
    """
    global _UNI_ID_MAP, _UNI_ID_MAP_WARNED
    if _UNI_ID_MAP:  # non-empty: keep
        return _UNI_ID_MAP

    for path in _map_paths():
        if not path.is_file():
            continue
        try:
            suffix = path.suffix.lower()
            if suffix == ".json":
                mp = _load_map_from_json(path)
            elif suffix == ".csv":
                mp = _load_map_from_csv(path)
            elif suffix in (".xlsx", ".xls"):
                mp = _load_map_from_xlsx(path)
            else:
                logger.warning("fino uni_id map unsupported suffix: %s", path)
                continue
            if not mp:
                continue
            _UNI_ID_MAP = mp
            logger.info("fino uni_id map loaded: %s (%d entries)", path, len(mp))
            return _UNI_ID_MAP
        except Exception as e:  # noqa: BLE001 — map is optional enrichment
            logger.warning("fino uni_id map failed to load %s: %s", path, e)

    _UNI_ID_MAP = {}
    if not _UNI_ID_MAP_WARNED:
        _UNI_ID_MAP_WARNED = True
        logger.warning(
            "fino uni_id map not found; detail.source will fall back to 机构#<uni_id>. "
            "Place mapping at backend/data/fino_uni_id_map.json "
            "(or set FINO_UNI_ID_MAP). Original run.py file: uni_id_ v1.1_20260119.xlsx"
        )
    return _UNI_ID_MAP


def reload_uni_id_map() -> dict[str, str]:
    """Force reload mapping (tests / after dropping a new file)."""
    global _UNI_ID_MAP, _UNI_ID_MAP_WARNED
    _UNI_ID_MAP = None
    _UNI_ID_MAP_WARNED = False
    return _load_uni_id_map()


def _is_fallback_source(source: Any, uni_id: str) -> bool:
    """True if source is empty or the placeholder 机构#<uni_id> / Unknown."""
    s = str(source or "").strip()
    if not s:
        return True
    if s in ("Unknown", "未知机构"):
        return True
    if uni_id and s == f"机构#{uni_id}":
        return True
    return False


def _enrich_detail_source(rows: list[dict]) -> list[dict]:
    """Attach source from uni_id map; fallback 机构#<uni_id>.

    Re-applies mapping even when a previous run left placeholder source values,
    so a newly added map file takes effect without waiting for cache TTL.
    """
    if not rows:
        return rows
    mp = _load_uni_id_map()
    out: list[dict] = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        r = dict(row)
        uid = str(r.get("uni_id") or "").strip()
        if not _is_fallback_source(r.get("source"), uid):
            out.append(r)
            continue
        if uid and uid in mp:
            r["source"] = mp[uid]
        elif uid:
            r["source"] = f"机构#{uid}"
        else:
            r["source"] = "未知机构"
        out.append(r)
    return out


def get_overview(
    report_type: str = "daily",
    start_date: str = "",
    end_date: str = "",
    product_codes: list[str] | None = None,
) -> list[dict]:
    """机构观点汇总.

    Args:
        report_type: daily / weekly
        start_date: 起始日 YYYYMMDD (空则用今天)
        end_date: 截止日 YYYYMMDD (空则用今天)
        product_codes: 品种代码列表 如 ["CU","RB"], 空则返回全量

    Returns:
        list[dict], 每条含 product_name / product_code / date /
        bull_count / neutral_count / bear_count /
        bull_percentage / neutral_percentage / bear_percentage /
        bull_views / neutral_views / bear_views /
        consensus_views / disagreement_views
    """
    rt = (report_type or "daily").strip()
    if rt not in ("daily", "weekly"):
        rt = "daily"
    sd = _norm_date(start_date) or time.strftime("%Y%m%d")
    ed = _norm_date(end_date) or sd
    codes = [c.strip().upper() for c in (product_codes or []) if c and c.strip()]

    appkey, appsecret = _creds()
    payload = {
        "report_type": rt,
        "start_date": sd,
        "end_date": ed,
        "product_code_list": codes,
        "appkey": appkey,
        "appsecret": appsecret,
    }
    key = f"fino:overview:{rt}:{sd}:{ed}:{','.join(codes)}"
    return _cached(
        key,
        lambda: _post(OVERVIEW_PATH, payload),
        valid=lambda v: isinstance(v, list) and len(v) > 0,
        ttl=600,
    )


def get_detail(
    report_type: str = "daily",
    start_date: str = "",
    end_date: str = "",
    product_codes: list[str] | None = None,
) -> list[dict]:
    """单品种机构观点明细 (逐条).

    上游字段: date / viewpoint / rating / detail / product_code / product_name / uni_id.
    本函数按本地映射表补全 source(机构名); 映射缺失时 source=\"机构#<uni_id>\".
    """
    rt = (report_type or "daily").strip()
    if rt not in ("daily", "weekly"):
        rt = "daily"
    sd = _norm_date(start_date) or time.strftime("%Y%m%d")
    ed = _norm_date(end_date) or sd
    codes = [c.strip().upper() for c in (product_codes or []) if c and c.strip()]

    appkey, appsecret = _creds()
    payload = {
        "report_type": rt,
        "start_date": sd,
        "end_date": ed,
        "product_code_list": codes,
        "appkey": appkey,
        "appsecret": appsecret,
    }
    # Cache RAW upstream rows; enrich on every return so map updates apply immediately.
    key = f"fino:detail:raw:{rt}:{sd}:{ed}:{','.join(codes)}"

    def _fetch() -> list[dict]:
        raw = _post(DETAIL_PATH, payload)
        rows = raw if isinstance(raw, list) else []
        return [r for r in rows if isinstance(r, dict)]

    cached = _cached(
        key,
        _fetch,
        valid=lambda v: isinstance(v, list) and len(v) > 0,
        ttl=600,
    )
    return _enrich_detail_source(cached if isinstance(cached, list) else [])
